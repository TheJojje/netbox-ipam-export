#!/opt/netbox/venv/bin/python3

from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows  # Lägg till denna import
import pandas as pd
import re
import ipaddress
import requests
import urllib3
from datetime import datetime

# Konfiguration
BASE_URL = "https://127.0.0.1/api/"
TOKEN = "b42fa7a7925a84e6835f54deeede67518ef12acc"
HEADERS = {"Authorization": f"Token {TOKEN}"}

# Hämta aktuellt datum och tid för filnamnet
DATE = datetime.now().strftime("%Y-%m-%d_%H_%M")

# Anpassa för att ignorera SSL-verifiering
session = requests.Session()
session.verify = False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Hämta data från NetBox API
def get_all_data(endpoint):
    results = []
    url = BASE_URL + endpoint
    while url:
        response = session.get(url, headers=HEADERS)
        response.raise_for_status()
        data = response.json()
        results.extend(data["results"])
        url = data.get("next")  # Hämtar nästa sida om den finns
    return results

# Exportera IPAM till Excel
def export_ipam():
    prefixes = get_all_data("ipam/prefixes/")
    ip_addresses = get_all_data("ipam/ip-addresses/")  # Hämta alla IP-adresser
    vlans = get_all_data("ipam/vlans/")  # Hämta alla VLANs
    ipam_writer = None  # Initiera Excel-skrivare

    try:
        ipam_writer = pd.ExcelWriter(f"IPAM_Export_{DATE}.xlsx", engine="openpyxl")

        # Innehållsförteckning
        # Skapa en dummy-flik om inga IP-adresser finns
        workbook = ipam_writer.book
        workbook.create_sheet("Dummy")  # Skapa en dummy-flik som alltid finns
        first_sheet = "Dummy"  # Använd dummy-fliken som den första

        # Skapa innehållsförteckning (TOC) flik
        toc_sheet = workbook.create_sheet("Innehållsförteckning", 0)  # Skapa fliken först
        toc_sheet.append(["Prefix", "Beskrivning", "VLAN ID", "VRF", "Roll", "Länk till prefix"])
        toc_sheet.column_dimensions['A'].width = 30
        toc_sheet.column_dimensions['B'].width = 40
        toc_sheet.column_dimensions['C'].width = 15
        toc_sheet.column_dimensions['D'].width = 20
        toc_sheet.column_dimensions['E'].width = 15
        toc_sheet.column_dimensions['F'].width = 30

        toc_sheet.row_dimensions[1].font = Font(bold=True)

        # Applicera bakgrund och svart ram på rubriker
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        for col in range(1, 7):  # Gäller alla kolumner i rubriken
            cell = toc_sheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="000000")
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for prefix in prefixes:
            # Kontrollera om prefix har en roll och om rollen är "Prefix"
            role = prefix.get("role")  # Kan vara None
            role_name = role["name"] if role else None

            if role_name == "Prefix":
                print(f"Hoppar över {prefix['prefix']} med rollen 'Prefix'")
                continue

            prefix_cidr = prefix["prefix"]
            prefix_network = ipaddress.IPv4Network(prefix_cidr, strict=False)

            # Hitta VLAN-ID
            vlan_id = None
            if "vlan" in prefix and prefix["vlan"]:
                vlan_id = prefix["vlan"].get("vid", "Inget VLAN")
            else:
                vlan_id = "N/A"

            # Hitta VRF
            vrf = None
            if "vrf" in prefix and prefix["vrf"]:
                vrf = prefix["vrf"].get("name", "Ingen VRF")
            else:
                vrf = "Ingen VRF"

            # Hämta prefixbeskrivning
            prefix_description = prefix.get("description", "Ingen beskrivning")

            # Ersätt otillåtna tecken för fliknamn och länkar
            sanitized_prefix_cidr = re.sub(r"[\/\\\?\*\[\]:]", "_", prefix_cidr)

            # Lägg till informationen i innehållsförteckningen
            toc_sheet.append([prefix_cidr, prefix_description, vlan_id, vrf, role_name, f"=HYPERLINK(\"#{sanitized_prefix_cidr}!A1\", \"Gå till prefix\")"])

            # Filtrera IP-adresser som ligger inom prefixets intervall
            ips_in_prefix = [
                ip for ip in ip_addresses
                if ipaddress.IPv4Address(ip["address"].split("/")[0]) in prefix_network
            ]

            print(f"IP-adresser exporterade för {prefix_cidr}: {len(ips_in_prefix)}")

        # Flikar för prefix
            # Skapa alltid fliken, även om det inte finns några IP-adresser
            if not ips_in_prefix:
                # Om inga IP-adresser finns, skapa en rad med meddelande
                ip_data = [{"Address": "Inga IP-adresser", "Description": ""}]
            else:
                # Formatera IP-data utan prefix
                ip_data = [
                    {
                        "Address": ip["address"].split("/")[0],  # Tar bort prefixdelen
                        "Description": ip.get("description", "")
                    }
                    for ip in ips_in_prefix
                ]

            # Skriv data till fliken
            df = pd.DataFrame(ip_data)
            df.to_excel(ipam_writer, sheet_name=sanitized_prefix_cidr, index=False, startrow=7)

            # Anpassa kolumnbredden och lägg till prefixinfo högst upp
            worksheet = workbook[sanitized_prefix_cidr]

            # Lägg till en rad för "Prefixinformation"
            worksheet.merge_cells("A1:B1")
            worksheet["A1"] = "Prefixinformation"
            worksheet["A1"].font = Font(bold=True)
            worksheet["A1"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

            # Lägg till prefixinformation
            header_data = [
                ["Prefix", prefix_cidr],
                ["VLAN ID", vlan_id],
                ["VRF", vrf],
                ["Roll", role_name if role_name else "Ingen roll"],
                ["Beskrivning", prefix_description],
            ]

            for row_index, (label, value) in enumerate(header_data, start=2):
                worksheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
                worksheet.cell(row=row_index, column=2, value=value)

            # Lägg till en ram runt informationen
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for row_index in range(2, len(header_data) + 2):
                for col_index in range(1, 3):
                    cell = worksheet.cell(row=row_index, column=col_index)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Justera kolumnbredden
            for column_cells in worksheet.columns:
                max_length = max(
                    (len(str(cell.value)) if cell.value else 0) for cell in column_cells
                )
                adjusted_width = max_length + 2  # Lägger till extra utrymme
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Lägg till länk tillbaka till innehållsförteckningen i B6
            worksheet.cell(row=7, column=2, value="Tillbaka till innehållsförteckning")
            worksheet.cell(row=7, column=2).font = Font(underline="single", color="0000FF")
            worksheet.cell(row=7, column=2).hyperlink = "#Innehållsförteckning!A1"  # Länk till TOC

            # Lägg till ett radavstånd innan IP-adresserna
            worksheet.insert_rows(8)

            # Gör rubriker för IP-adresserna och beskrivning fetmarkerade med grå bakgrund
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            worksheet.cell(row=9, column=1).font = Font(bold=True)
            worksheet.cell(row=9, column=1).fill = header_fill
            worksheet.cell(row=9, column=2).font = Font(bold=True)
            worksheet.cell(row=9, column=2).fill = header_fill

            # Sätt första fliken som aktiv
            if first_sheet == "Dummy":
                first_sheet = sanitized_prefix_cidr

        # Om det finns en flik, sätt den som aktiv (Sätt "Innehållsförteckning" som aktiv)
        workbook.active = workbook.sheetnames.index("Innehållsförteckning")

        # Ta bort dummy-fliken om den är kvar
        if "Dummy" in workbook.sheetnames:
            del workbook["Dummy"]

        # Spara filen genom att använda rätt metod
        workbook.save(f"IPAM_Export_{DATE}.xlsx")  # Rätt metod för att spara boken

    except Exception as e:
        print(f"Ett fel uppstod: {e}")
    finally:
        # Stäng skrivaren om den har skapats
        if ipam_writer:
            ipam_writer._save()

# Kör exporterna
export_ipam()
